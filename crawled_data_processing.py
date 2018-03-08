# -*- coding: utf-8 -*-

import os,re
import os.path
import datetime as dt
import jieba
import jieba.posseg

import sys
import nltk

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter 

import threading

reload(sys)
sys.setdefaultencoding('utf-8')

# pattern = re.compile(r'.*\.xlsx') #编译个正则表达式
# listfiles = []
# target_files = []
# tempgch = []
# fenci_stopwords = []

def getgch(pattern,listfiles):
    target_files = []
    tempgch = []
    for files in listfiles:   #把传入的文件列表逐个取出用'\\'分开后写入list
        # global pattern
        # global target_files
        # global tempgch
        a = files.split('\\')
        for i in a:         #将list逐个取出和正则表达式匹配
            match = pattern.match(i)
            if match:
                tempgch.append(match.group()) #把匹配的结果放到临时list中
    return list(set(tempgch))      #把临时list去重

# 遍历目标路径的所有正则规定的文件       
def walkdir(pattern,srcdir):    #遍历目录
    listfiles = []
    srcdir = os.path.abspath(srcdir)
    for parent,dirs,files in os.walk(srcdir):
        for file in files:
            infile = os.path.join(parent,file)
            listfiles.append(infile)
    #print listfiles
    return getgch(pattern,listfiles)   #返回文件列表，这里似乎更应该直接返回dirs这个列表即可

 #对目标文件的文本进行分词
# def fenci(text):

def load_stopwords(pattern, filepath):
    fenci_stopwords = []
    stopfiles = walkdir(pattern, filepath)
    print stopfiles
    for stopf in stopfiles: 
        fstop = open(stopf)
        print stopf
        for eachWord in fstop:
            try:
                fenci_stopwords.append(eachWord.strip().decode('utf-8','ignore'))
            except:
                continue
    return list(set(fenci_stopwords))

def iterate_files_assemble_fenci(fenci_stopwords, listfiles):
    threads = []
    for f in listfiles:
        t = threading.Thread(target=thread_core,args=(fenci_stopwords,f))
        t.setDaemon(True)
        t.start()
        threads.append(t)
    for t in threads:
        t.join()

    print "all over"
        # cut_wl = []
        # wb = load_workbook(f)
        # new_wb = Workbook()
        # new_ws = new_wb.active
        # sheetnames = wb.get_sheet_names()
        # sheet = wb.get_sheet_by_name(sheetnames[0])
        # print ">>>>>>> %s" % f
        # row_length = sheet.max_row
        # col_length = sheet.max_column
        # print "row number: %d" % row_length
        # print "col number: %d" % col_length
        # for line in range(2,row_length+1):
        #     job_id = sheet.cell(row=line,column=1).value
        #     job_desc = sheet.cell(row=line,column=2).value
        #     seg = jieba.posseg.cut(job_desc)
        #     for item in seg:
        #         if item.word not in fenci_stopwords:
        #             last_number = len(cut_wl) + 1
        #             cut_wl.append((last_number, item.word, item.flag, job_id))
        #             new_ws.append([last_number, item.word, item.flag, job_id])
        # new_f = f.split(".xlsx")[0] + "_fenci.xlsx"
        # new_wb.save(new_f)

def thread_core(fenci_stopwords, targetfile):
    new_f = targetfile.split(".xlsx")[0] + "_fenci.xlsx"
    if os.path.exists(new_f):
        print "file existed, skip!"
        return true
    cut_wl = []
    wb = load_workbook(targetfile)
    new_wb = Workbook()
    new_ws = new_wb.active
    sheetnames = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(sheetnames[0])
    print ">>>>>>> %s" % targetfile
    row_length = sheet.max_row
    col_length = sheet.max_column
    print "row number: %d" % row_length
    print "col number: %d" % col_length
    for line in range(2,row_length+1):
        job_id = sheet.cell(row=line,column=1).value
        job_desc = sheet.cell(row=line,column=6).value
        if job_desc is not None:
            # seg = jieba.posseg.cut(job_desc.encode('utf-8', 'ignore'))
            # for item in seg:
            #     if item.word not in fenci_stopwords:
            #         try:
            #             last_number = len(cut_wl) + 1
            #             cut_wl.append((last_number, item.word, item.flag, job_id))
            #             new_ws.append([last_number, item.word.encode('utf-8', 'ignore'), item.flag, job_id])
            #         except:
            #             continue
            tokens = nltk.word_tokenize(job_desc)
            seg = nltk.pos_tag(tokens)           
            for item in seg:
                if item[0] not in fenci_stopwords:
                    try:
                        last_number = len(cut_wl) + 1
                        cut_wl.append((last_number, item[0], item[1], job_id))
                        new_ws.append([last_number, item[0].encode('utf-8', 'ignore'), item[1], job_id])
                    except:
                        continue

    new_wb.save(new_f)
    print "file: %s, finished!" % targetfile

        # print sheet.cell(row=2,column=10).value
        # seg = jieba.posseg.cut(sheet.cell(row=2,column=10).value)
        # l = []
        # for i in seg:
        #     if i.word not in fenci_stopwords:
        #         l.append((i.word, i.flag))
        # for element in l:
        #     print element[0],element[1]            
# #对获取的分词输出
# def output_fenci(target_fenci_arr):

if __name__ == "__main__":    
    srcdir = str(raw_input("please input the dirs:\n"))
    #遍历所有目标文件
    target_files = walkdir(re.compile(r'.*desc\.xlsx'),srcdir)
    target_stopwords = load_stopwords(re.compile(r'.*\.txt'), "/Users/freekick/Workspace/qiancai/rencai_bj/src/stopwords/")
    iterate_files_assemble_fenci(target_stopwords, target_files)
